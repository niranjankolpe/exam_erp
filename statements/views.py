from django.shortcuts import render
from django.http import HttpResponse

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from io import BytesIO

from reportlab.lib.pagesizes import landscape, legal 
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from reportlab.platypus import Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle

from django.conf import settings
from statements.models import *

import time

# Create your views here.
def populate_db():
    with open('static/raw_data.txt', 'r') as file:
        lines = file.readlines()
    
    for line in lines:
        try:
            data = line.strip()
            record = University(name=data)
            record.save()
        except Exception as e:
            print(e)
            continue
    return

def statement(request):
    # populate_db()
    return render(request, "select_statement_type.html")

def labstaff(request):
    universities = University.objects.all()
    colleges = College.objects.all()
    departments = Department.objects.all()
    subjects = Subject.objects.all()
    teachers = Teacher.objects.all()
    context = {'universities': universities, 'colleges': colleges, 'departments': departments, 'subjects': subjects, 'teachers':teachers}
    if request.method == 'POST':
        input_str_list = ["college_name", "exam_start_month", "exam_start_year", "exam_end_month", "exam_end_year", "department"]
        input_int_list = ["subject_count", "rate_expert_PandE", "rate_lab_assist_prep", "rate_lab_assist_exam", "rate_tech_assist_prep", "rate_tech_assist_exam", "rate_peon_prep", "rate_peon_exam", "rate_of_electricity"]
        data = dict()
        for i in input_str_list:
            data.update({i: request.POST[i]})
        data.update({'time_period': data["exam_start_month"] + " " + data["exam_start_year"] + " - " + data["exam_end_month"] + " " + data["exam_end_year"]})
        for i in input_int_list:
            data.update({i: int(request.POST[i])})
        return generate_invoice(request, data) if request.POST['excel_pdf'] == "excel" else labstaff_pdf(request)
    return render(request, "statement_of_labstaff_form1.html", context)

def generate_invoice(request, data):
    wb = openpyxl.Workbook()

    # Lab Staff Excel Sheet
    college_name = data['college_name']
    exam_start_month = data["exam_start_month"]
    exam_start_year = data["exam_start_year"]
    exam_end_month = data["exam_end_month"]
    exam_end_year = data["exam_end_year"]

    time_period = exam_start_month + " " + exam_start_year + " - " + exam_end_month + " " + exam_end_year
    department = data['department']
    subject_count = int(data['subject_count'])
    rate_expert_PandE = int(data['rate_expert_PandE'])
    rate_lab_assist_prep = int(data['rate_lab_assist_prep'])
    rate_lab_assist_exam = int(data['rate_lab_assist_exam'])
    rate_tech_assist_prep = int(data['rate_tech_assist_prep'])
    rate_tech_assist_exam = int(data['rate_tech_assist_exam'])
    rate_peon_prep = int(data['rate_peon_prep'])
    rate_peon_exam = int(data['rate_peon_exam'])

    wb = openpyxl.Workbook()

    # Statement of Lab Staff
    statement_of_lab_staff = wb.active
    statement_of_lab_staff.title = "STATEMENT_OF_LABSTAFF"

    header_data1 = [
        ['NAME OF THE COLLEGE : ' + college_name.upper()],
        [f'FACULTY OF ENGINEERING ({department.upper()})'],
        ['STATEMENT SHOWING THE STAFF USED FOR PRACTICAL EXAMINATION HELD IN THE COLLEGE'],
        [f'For {time_period}'],
        ['','','','','','','',"Total No. of Supporting Staff used",'','','','','','','','','','',"Payment to Supporting Staff As per rate  per batch + preparation & Cleaning"],
        ["", "", "", "", "", "", "", "Expert", "", "Lab", "", "", "Tech", "", "", "Peon", "", "", "Expert", "", "Lab", "", "Tech", "", "Peon", "", "Total", ""],
        ["Year & Course", "Subject of Exam", "PR/OR/TW", "No. of Students", "Dates of conduct of Exam", 
        "Days of Prep & Clng", "No. of Batches", 
        "No. of Expert Asstt", "Rate for P & E", 
        "Lab Asstt", "Rate for Prep", "Rate for Exam", 
        "Tech Asstt", "Rate for Prep", "Rate for Exam", 
        "Peon", "Rate for Prep", "Rate for Exam", 
        "Expert Asstt exam", "Amount for Prep & Clng", 
        "Lab Asstt exam", "Amount for Prep & Clng", 
        "Tech Asstt exam", "Amount for Prep & Clng", 
        "Peon exam", "Amount for Prep & Clng", 
        "Total for Prep & Clean", "Rem exam", "Total rem"]
    ]

    for row in header_data1:
        statement_of_lab_staff.append(row)

    statement_of_lab_staff.merge_cells('A1:AC1')
    statement_of_lab_staff.merge_cells('A2:AC2')
    statement_of_lab_staff.merge_cells('A3:AC3')
    statement_of_lab_staff.merge_cells('A4:AC4')
    statement_of_lab_staff.merge_cells('H5:R5')
    statement_of_lab_staff.merge_cells('S5:AB5')
    statement_of_lab_staff.merge_cells('H6:I6')
    statement_of_lab_staff.merge_cells('J6:L6')
    statement_of_lab_staff.merge_cells('M6:O6')
    statement_of_lab_staff.merge_cells('P6:R6')
    statement_of_lab_staff.merge_cells('S6:T6')
    statement_of_lab_staff.merge_cells('U6:V6')
    statement_of_lab_staff.merge_cells('W6:X6')
    statement_of_lab_staff.merge_cells('Y6:Z6')

    for i in range(1, subject_count+1):
        academic_year = request.POST['subject'+ str(i) +'_academic_year']
        subject_name = request.POST['subject'+ str(i)]
        pr_or_tw = request.POST['subject'+ str(i) +'_pr_or_tw']
        student_count = int(request.POST['subject'+ str(i) +'_student_count'])
        date_of_exam = request.POST['subject'+ str(i) +'_date_of_conduct']
        days_of_prep_and_clean = int(request.POST['subject'+ str(i) +'_days_of_preparation'])
        batch_count = int(request.POST['subject'+ str(i) +'_batch_count'])
        expert_count = int(request.POST['subject'+ str(i) +'_expert_count'])
        lab_assist_count = int(request.POST['subject'+ str(i) +'_lab_assistant_count'])   
        tech_assist_count = int(request.POST['subject'+ str(i) +'_tech_assistant_count'])
        peon_count = int(request.POST['subject'+ str(i) +'_peon_count'])

        payment_expert_exam = expert_count * rate_expert_PandE * batch_count
        payment_expert_prep_and_clean = expert_count * rate_expert_PandE * days_of_prep_and_clean
        payment_lab_assist_exam = lab_assist_count * rate_lab_assist_exam * batch_count
        payment_lab_assist_prep_and_clean = lab_assist_count * rate_lab_assist_prep * days_of_prep_and_clean
        payment_tech_assist_exam = tech_assist_count * rate_tech_assist_exam * batch_count
        payment_tech_assist_prep_and_clean = tech_assist_count * rate_tech_assist_prep * days_of_prep_and_clean
        payment_peon_exam = peon_count * rate_peon_exam * batch_count
        payment_peon_prep_and_clean = peon_count * rate_peon_prep * days_of_prep_and_clean

        payment_total_prep_and_clean = payment_expert_prep_and_clean + payment_lab_assist_prep_and_clean + payment_tech_assist_prep_and_clean + payment_peon_prep_and_clean
        payment_total_rem_exam = payment_expert_exam + payment_lab_assist_exam + payment_tech_assist_exam + payment_peon_exam
        total_rem = payment_total_prep_and_clean + payment_total_rem_exam

        # Append data
        statement_of_lab_staff.append([
            academic_year + " " + department,
            subject_name,
            pr_or_tw,
            student_count,
            date_of_exam,
            days_of_prep_and_clean,
            batch_count,
            expert_count,
            rate_expert_PandE,
            lab_assist_count,
            rate_lab_assist_prep,
            rate_lab_assist_exam,
            tech_assist_count,
            rate_tech_assist_prep,
            rate_tech_assist_exam,
            peon_count,
            rate_peon_prep,
            rate_peon_exam,
            payment_expert_exam,
            payment_expert_prep_and_clean,
            payment_lab_assist_exam,
            payment_lab_assist_prep_and_clean,
            payment_tech_assist_exam,
            payment_tech_assist_prep_and_clean,
            payment_peon_exam,
            payment_peon_prep_and_clean,
            payment_total_prep_and_clean,
            payment_total_rem_exam,
            total_rem
        ])

    header_font = Font(bold=True, size=10, name="Times New Roman", color = "002060")
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True, shrink_to_fit=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in statement_of_lab_staff.iter_rows(min_row=1, min_col=1, max_col=29):
        for cell in row:
            cell.font = header_font
            cell.alignment = center_alignment

    for row in statement_of_lab_staff.iter_rows(min_row=4, min_col=1, max_col=29):
        for cell in row:        
            cell.border = thin_border

    statement_of_lab_staff.column_dimensions["B"].width = 12


    # Statement of Internal/External Bill Excel Sheet
    statement_of_internal_external_bill = wb.create_sheet(title="int_ext_bill")
    header_data2 = [
        ['NAME OF THE COLLEGE : ' + college_name.upper()],
        [f'STATEMENT SHOWING  OF REMUNERATION PAID TO EXAMINERS FOR CONDUCTING PRACTICAL/ ORAL EXAM  {time_period} Examination'],
        
        ["Sr. No.","Name of the Examiner", " CATEGORY (INTERNAL / EXTERNAL)",'Subject',
        'Class','Date', "OR/ PR/ TW", "No. of Students", "Rate", 
        "Payable Rate", "Total=  (Payable Rate*No. of Students)", 'Remark']
    ]

    # Fill the first header row and merge cells
    for row in header_data2:
        statement_of_internal_external_bill.append(row)

    # Merging cells for first row as per the requirement
    statement_of_internal_external_bill.merge_cells('A1:L1')
    statement_of_internal_external_bill.merge_cells('A2:L2')

    # Append data to the sheet based on subject count
    for i in range(1, subject_count+1):
        subject_name = request.POST['subject'+ str(i)]
        examiner_name = request.POST['subject'+ str(i)+'_examiner_name']
        category = request.POST['subject'+ str(i) +'_category']
        academic_year = request.POST['subject'+ str(i) +'_academic_year']
        date_of_exam = request.POST['subject'+ str(i) +'_date_of_conduct']
        pr_or_tw = request.POST['subject'+ str(i) +'_pr_or_tw']
        student_count = int(request.POST['subject'+ str(i) +'_student_count'])
        rate = int(request.POST['subject'+ str(i) +'_rate'])
        payable_rate = int(request.POST['subject'+ str(i) +'_payable_rate'])
        
        total = student_count * payable_rate

        # Append data
        statement_of_internal_external_bill.append([
            i,
            examiner_name,
            category,
            subject_name,
            academic_year,
            date_of_exam,
            pr_or_tw,
            student_count,
            rate,
            payable_rate,
            total
        ])

    # Apply bold and alignment for header rows
    header_font = Font(bold=True, size=10, name="Times New Roman", color = "002060")
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True, shrink_to_fit=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in statement_of_internal_external_bill.iter_rows(min_row=1, min_col=1, max_col=12):
        for cell in row:
            cell.font = header_font
            cell.alignment = center_alignment

    for row in statement_of_internal_external_bill.iter_rows(min_row=1, min_col=1, max_col=12):
        for cell in row:        
            cell.border = thin_border

    for row in statement_of_internal_external_bill.iter_rows(min_row=3, max_row=3, min_col=3, max_col=3):
        for cell in row:
            cell.font = Font(size=8, name="Times New Roman", color = "002060")


    statement_of_internal_external_bill.column_dimensions["A"].width = 5
    statement_of_internal_external_bill.column_dimensions["B"].width = 20
    statement_of_internal_external_bill.column_dimensions["D"].width = 25


    # Electricity Bill
    electricity_bill = wb.create_sheet(title="electricity_bill")
    header_data3 = [
        ['NAME OF THE COLLEGE : ' + college_name.upper()],
        ['Statement Showing Claim for use of Computation Facilities, Electricity and allied'],
        [f'EXAM :SE/TE/BE/ME {department} {time_period}'],
        [""],
        ['Sr.No.','Date of Exam','Class','PR/OR/TW/Seminar/Project/Theory',
        'No. of Students', 'Rate','Amount'],
    ]

    # Fill the first header row and merge cells
    for row in header_data3:
        electricity_bill.append(row)

    # Merging cells for first row as per the requirement
    electricity_bill.merge_cells('A1:G1')
    electricity_bill.merge_cells('A2:G2')
    electricity_bill.merge_cells('A3:G3')
    electricity_bill.merge_cells('A4:G4')

    # Append data to the sheet based on subject count
    for i in range(1, subject_count+1):
        subject = request.POST[f"subject{i}"]
        academic_year = request.POST[f"subject{i}_academic_year"]
        exam_date = request.POST[f"subject{i}_date_of_conduct"]
        student_count = int(request.POST[f"subject{i}_student_count"])
        rate_of_electricity = data["rate_of_electricity"]
        
        amount = rate_of_electricity * student_count

        # Append data
        electricity_bill.append([
            i,
            exam_date,
            academic_year,
            subject,
            student_count,
            rate_of_electricity,
            amount
        ])

    # Apply bold and alignment for header rows
    header_font = Font(bold=True, size=10, name="Times New Roman", color = "002060")
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in electricity_bill.iter_rows(min_row=1, min_col=1, max_col=7):
        for cell in row:
            cell.font = header_font
            cell.alignment = center_alignment

    for row in electricity_bill.iter_rows(min_row=1, min_col=1, max_col=7):
        for cell in row:        
            cell.border = thin_border

    electricity_bill.column_dimensions["A"].width = 7
    electricity_bill.column_dimensions["B"].width = 10
    electricity_bill.column_dimensions["C"].width = 8
    electricity_bill.column_dimensions["D"].width = 40
    electricity_bill.column_dimensions["E"].width = 10
    electricity_bill.column_dimensions["F"].width = 6
    electricity_bill.column_dimensions["G"].width = 7
    

    # Saving the file
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    response = HttpResponse(file_stream, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="invoice.xlsx"'
    return response

def labstaff_pdf(request):
    if request.method == 'POST':
        college_name = request.POST['college_name']
        # university = request.POST['university']
        exam_start_month = request.POST["exam_start_month"]
        exam_start_year = request.POST["exam_start_year"]
        exam_end_month = request.POST["exam_end_month"]
        exam_end_year = request.POST["exam_end_year"]

        time_period = exam_start_month + " " + exam_start_year + " - " + exam_end_month + " " + exam_end_year
        
        # time_period = request.POST['time_period']
        department = request.POST['department']
        subject_count = int(request.POST['subject_count'])
        rate_expert_PandE = int(request.POST['rate_expert_PandE'])
        rate_lab_assist_prep = int(request.POST['rate_lab_assist_prep'])
        rate_lab_assist_exam = int(request.POST['rate_lab_assist_exam'])
        rate_tech_assist_prep = int(request.POST['rate_tech_assist_prep'])
        rate_tech_assist_exam = int(request.POST['rate_tech_assist_exam'])
        rate_peon_prep = int(request.POST['rate_peon_prep'])
        rate_peon_exam = int(request.POST['rate_peon_exam'])

        # Define headers for the table with merged cells

        paragraphStyle_header = ParagraphStyle(name='CustomStyle',fontSize=12,fontName='Times-Bold',alignment=1)
        data = [
            [Paragraph(f'NAME OF THE COLLEGE : {college_name.upper()}',paragraphStyle_header)],
            [Paragraph(f'FACULTY OF ENGINEERING ({department.upper()})',paragraphStyle_header)],
            [Paragraph('STATEMENT SHOWING THE STAFF USED FOR PRACTICAL EXAMINATION HELD IN THE COLLEGE',paragraphStyle_header)],
            [Paragraph(f'For {time_period}',paragraphStyle_header)],
            ['','','','','','','',"Total No. of Supporting Staff used",'','','','','','','','','','',"Payment to Supporting Staff As per rate  per batch + preparation & Cleaning"],
            ["", "", "", "", "", "", "", "Expert", "", "Lab", "", "", "Tech", "", "", "Peon", "", "", "Expert", "", "Lab", "", "Tech", "", "Peon", "", "Total", ""],
            [Paragraph("Year & Course",ParagraphStyle(name='CustomStyle',fontSize=8,)), 
            Paragraph("Subject of Exam",ParagraphStyle(name='CustomStyle',fontSize=8,)), 
            Paragraph("PR/ OR/ TW",ParagraphStyle(name='CustomStyle',fontSize=8,)), 
            Paragraph("No. of Students",ParagraphStyle(name='CustomStyle',fontSize=8,)), 
            Paragraph("Dates of conduct of Exam",ParagraphStyle(name='CustomStyle',fontSize=8,)), 
            Paragraph("Days of Prep & Clng",ParagraphStyle(name='CustomStyle',fontSize=8,)), 
            Paragraph("No. of Batches",ParagraphStyle(name='CustomStyle',fontSize=8,)), 
            Paragraph("No. of Expert Asstt",ParagraphStyle(name='CustomStyle',fontSize=8,)), 
            Paragraph("Rate for P & E",ParagraphStyle(name='CustomStyle',fontSize=8,)), 
            Paragraph("Lab Asstt",ParagraphStyle(name='CustomStyle',fontSize=8,)), 
            Paragraph("Rate for Prep",ParagraphStyle(name='CustomStyle',fontSize=8,)), 
            Paragraph("Rate for Exam",ParagraphStyle(name='CustomStyle',fontSize=8,)), 
            Paragraph("Tech Asstt",ParagraphStyle(name='CustomStyle',fontSize=8,)), 
            Paragraph("Rate for Prep",ParagraphStyle(name='CustomStyle',fontSize=8,)), 
            Paragraph("Rate for Exam",ParagraphStyle(name='CustomStyle',fontSize=8,)), 
            Paragraph("Peon",ParagraphStyle(name='CustomStyle',fontSize=8,)), 
            Paragraph("Rate for Prep",ParagraphStyle(name='CustomStyle',fontSize=8,)), 
            Paragraph("Rate for Exam",ParagraphStyle(name='CustomStyle',fontSize=8,)), 
            Paragraph("Expert Asstt exam",ParagraphStyle(name='CustomStyle',fontSize=8,)), 
            Paragraph("Amount for Prep & Clng",ParagraphStyle(name='CustomStyle',fontSize=8,)), 
            Paragraph("Lab Asstt exam",ParagraphStyle(name='CustomStyle',fontSize=8,)), 
            Paragraph("Amount for Prep & Clng",ParagraphStyle(name='CustomStyle',fontSize=8,)), 
            Paragraph("Tech Asstt exam",ParagraphStyle(name='CustomStyle',fontSize=8,)), 
            Paragraph("Amount for Prep & Clng",ParagraphStyle(name='CustomStyle',fontSize=8,)), 
            Paragraph("Peon exam",ParagraphStyle(name='CustomStyle',fontSize=8,)), 
            Paragraph("Amount for Prep & Clng",ParagraphStyle(name='CustomStyle',fontSize=8,)), 
            Paragraph("Total for Prep & Clean",ParagraphStyle(name='CustomStyle',fontSize=8,)), 
            Paragraph("Rem exam",ParagraphStyle(name='CustomStyle',fontSize=8,)), 
            Paragraph("Total rem",ParagraphStyle(name='CustomStyle',fontSize=8,))]
        ]

        # Append data to the sheet based on subject count
        for i in range(1, subject_count+1):
            academic_year = request.POST['subject'+ str(i) +'_academic_year']
            subject_name = request.POST['subject'+ str(i)]
            pr_or_tw = request.POST['subject'+ str(i) +'_pr_or_tw']
            student_count = int(request.POST['subject'+ str(i) +'_student_count'])
            date_of_exam = request.POST['subject'+ str(i) +'_date_of_conduct']
            days_of_prep_and_clean = int(request.POST['subject'+ str(i) +'_days_of_preparation'])
            batch_count = int(request.POST['subject'+ str(i) +'_batch_count'])
            expert_count = int(request.POST['subject'+ str(i) +'_expert_count'])
            lab_assist_count = int(request.POST['subject'+ str(i) +'_lab_assistant_count'])   
            tech_assist_count = int(request.POST['subject'+ str(i) +'_tech_assistant_count'])
            peon_count = int(request.POST['subject'+ str(i) +'_peon_count'])

            payment_expert_exam = expert_count * rate_expert_PandE * batch_count
            payment_expert_prep_and_clean = expert_count * rate_expert_PandE * days_of_prep_and_clean
            payment_lab_assist_exam = lab_assist_count * rate_lab_assist_exam * batch_count
            payment_lab_assist_prep_and_clean = lab_assist_count * rate_lab_assist_prep * days_of_prep_and_clean
            payment_tech_assist_exam = tech_assist_count * rate_tech_assist_exam * batch_count
            payment_tech_assist_prep_and_clean = tech_assist_count * rate_tech_assist_prep * days_of_prep_and_clean
            payment_peon_exam = peon_count * rate_peon_exam * batch_count
            payment_peon_prep_and_clean = peon_count * rate_peon_prep * days_of_prep_and_clean

            payment_total_prep_and_clean = payment_expert_prep_and_clean + payment_lab_assist_prep_and_clean + payment_tech_assist_prep_and_clean + payment_peon_prep_and_clean
            payment_total_rem_exam = payment_expert_exam + payment_lab_assist_exam + payment_tech_assist_exam + payment_peon_exam
            total_rem = payment_total_prep_and_clean + payment_total_rem_exam

            data.append([
                Paragraph(academic_year + " " + department,ParagraphStyle(name='CustomStyle',fontSize=8,)),
                Paragraph(subject_name, ParagraphStyle(name='CustomStyle',fontSize=8,)),
                pr_or_tw,
                student_count,
                date_of_exam,
                days_of_prep_and_clean,
                batch_count,
                expert_count,
                rate_expert_PandE,
                lab_assist_count,
                rate_lab_assist_prep,
                rate_lab_assist_exam,
                tech_assist_count,
                rate_tech_assist_prep,
                rate_tech_assist_exam,
                peon_count,
                rate_peon_prep,
                rate_peon_exam,
                payment_expert_exam,
                payment_expert_prep_and_clean,
                payment_lab_assist_exam,
                payment_lab_assist_prep_and_clean,
                payment_tech_assist_exam,
                payment_tech_assist_prep_and_clean,
                payment_peon_exam,
                payment_peon_prep_and_clean,
                payment_total_prep_and_clean,
                payment_total_rem_exam,
                total_rem
            ])


    # Create the HttpResponse object with the appropriate PDF headers
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="STATEMENT_OF_LABSTAFF.pdf"'

    # Create the PDF object, using the HttpResponse object as its "file."
    pdf = canvas.Canvas(response, pagesize=landscape(legal))

    # Set the title
    pdf.setTitle("Exam Payments Report")
    
    col_widths = [50,70,30,30,40,30,30,30,30,30,30,30,30,40,20,30,30,30,35,30,35,30,35,30,35,30,30,30,]

    # Create the table
    table = Table(data,colWidths=col_widths)
    table.setStyle(TableStyle([
        ('SPAN', (0,0), (-1, 0)),  
        ('SPAN', (0, 1), (-1, 1)),
        ('SPAN', (0, 2), (-1, 2)),
        ('SPAN', (0, 3), (-1, 3)),
        ('SPAN', (7, 4), (17, 4)),
        ('SPAN', (18, 4), (27, 4)),

        ('SPAN', (7, 5), (8, 5)),  # Merge "Expert" header
        ('SPAN', (9, 5), (11, 5)),  # Merge "Lab" header
        ('SPAN', (12, 5), (14, 5)),  # Merge "Tech" header
        ('SPAN', (15, 5), (17, 5)),  # Merge "Peon" header
        ('SPAN', (18, 5), (19, 5)),  # Merge "Expert" header for second section
        ('SPAN', (20, 5), (21, 5)),  # Merge "Lab" header for second section
        ('SPAN', (22, 5), (23, 5)),  # Merge "Tech" header for second section
        ('SPAN', (24, 5), (25, 5)),  # Merge "Peon" header for second section
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 5), (-1, -1), 'Times-Roman'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('WORDWRAP', (0, 0), (-1, -1)),  # Enable text wrapping
        ('FONTSIZE', (0, 0), (-1, -1), 8),  # Reduce font size if necessary
        ('LEFTPADDING', (0, 0), (-1, -1), 1),  # Reduce left padding
        ('RIGHTPADDING', (0, 0), (-1, -1), 1),  # Reduce right padding
    ]))

    page_width, page_height = landscape(legal)

    # Draw the table on the PDF
    table_width, table_height = table.wrapOn(pdf, page_width-20, page_height)
    table.drawOn(pdf, 10, page_height - table_height - 20)


    # Close the PDF object cleanly
    pdf.showPage()
    pdf.save()

    return response

def internal_external_bill(request):
    if request.method == 'POST':
        return internal_external_bill_excel(request) if request.POST['excel_pdf'] == "excel" else internal_external_bill_pdf(request) 
    return render(request, "internal_external_bill.html")

def internal_external_bill_excel(request, data, wb):
    
    college_name = data['college_name']
    time_period = data['time_period']
    subject_count = data['subject_count']

    ws = wb.create_sheet("statement_of_internal_external_bill")

    # Define headers for the table with merged cells
    header_data = [
        ['NAME OF THE COLLEGE : ' + college_name.upper()],
        [f'STATEMENT SHOWING  OF REMUNERATION PAID TO EXAMINERS FOR CONDUCTING PRACTICAL/ ORAL EXAM  {time_period} Examination'],
        
        ["Sr. No.","Name of the Examiner", " CATEGORY (INTERNAL / EXTERNAL)",'Subject',
        'Class','Date', "OR/ PR/ TW", "No. of Students", "Rate", 
        "Payable Rate", "Total=  (Payable Rate*No. of Students)", 'Remark']
    ]

    # Fill the first header row and merge cells
    for row in header_data:
        ws.append(row)

    # Merging cells for first row as per the requirement
    ws.merge_cells('A1:L1')
    ws.merge_cells('A2:L2')

    # Append data to the sheet based on subject count
    for i in range(1, subject_count+1):
        subject_name = request.POST['subject'+ str(i)]
        examiner_name = request.POST['subject'+ str(i)+'_examiner_name']
        category = request.POST['subject'+ str(i) +'_category']
        academic_year = request.POST['subject'+ str(i) +'_academic_year']
        date_of_exam = request.POST['subject'+ str(i) +'_date_of_conduct']
        pr_or_tw = request.POST['subject'+ str(i) +'_pr_or_tw']
        student_count = int(request.POST['subject'+ str(i) +'_student_count'])
        rate = int(request.POST['subject'+ str(i) +'_rate'])
        payable_rate = int(request.POST['subject'+ str(i) +'_payable_rate'])
        
        total = student_count * payable_rate

        # Append data
        ws.append([
            i,
            examiner_name,
            category,
            subject_name,
            academic_year,
            date_of_exam,
            pr_or_tw,
            student_count,
            rate,
            payable_rate,
            total
        ])
    
    # Apply bold and alignment for header rows
    header_font = Font(bold=True, size=10, name="Times New Roman", color = "002060")
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True, shrink_to_fit=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in ws.iter_rows(min_row=1, min_col=1, max_col=12):
        for cell in row:
            cell.font = header_font
            cell.alignment = center_alignment

    for row in ws.iter_rows(min_row=1, min_col=1, max_col=12):
        for cell in row:        
            cell.border = thin_border

    for row in ws.iter_rows(min_row=3, max_row=3, min_col=3, max_col=3):
        for cell in row:
            cell.font = Font(size=8, name="Times New Roman", color = "002060")


    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["D"].width = 25

    # Save the workbook to a BytesIO object (in-memory stream)
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)  # Ensure you reset the stream position

    # Create the response as an Excel file
    # response = HttpResponse(file_stream, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    # response['Content-Disposition'] = 'attachment; filename="statement_of_internal_external_bill.xlsx"'
    return wb

def internal_external_bill_pdf(request):
    if request.method == 'POST':
        college_name = request.POST['college_name']
        university = request.POST['university']
        time_period = request.POST['time_period']
        subject_count = int(request.POST['subject_count'])

        # Define headers for the table with merged cells

        paragraphStyle_header = ParagraphStyle(name='CustomStyle',fontSize=12,fontName='Times-Bold',alignment=1)
        data = [
            [Paragraph(f'NAME OF THE COLLEGE : {college_name.upper()}',paragraphStyle_header)],
            [Paragraph(f'STATEMENT SHOWING  OF REMUNERATION PAID TO EXAMINERS FOR CONDUCTING PRACTICAL/ ORAL EXAM  {time_period} Examination',paragraphStyle_header)],
            
            [Paragraph("Sr. No.",ParagraphStyle(name='CustomStyle',fontSize=10,alignment=1)),
            Paragraph("Name of the Examiner", ParagraphStyle(name='CustomStyle',fontSize=10,alignment=1)),
            Paragraph(" CATEGORY (INTERNAL / EXTERNAL)",ParagraphStyle(name='CustomStyle',fontSize=9,alignment=1)),
            'Subject','Class','Date', 
            Paragraph("OR/ PR/ TW", ParagraphStyle(name='CustomStyle',fontSize=10,alignment=1)),
            Paragraph("No. of Students", ParagraphStyle(name='CustomStyle',fontSize=9,alignment=1,)),
            "Rate", 
            Paragraph("Payable Rate", ParagraphStyle(name='CustomStyle',fontSize=10,alignment=1,)),
            Paragraph("Total=  (Payable Rate * No. of Students)", ParagraphStyle(name='CustomStyle',fontSize=9,alignment=1,)),
            'Remark']
        ]

        # Append data to the sheet based on subject count
        for i in range(1, subject_count+1):
            subject_name = request.POST['subject'+ str(i)]
            examiner_name = request.POST['subject'+ str(i)+'_examiner_name']
            category = request.POST['subject'+ str(i) +'_category']
            academic_year = request.POST['subject'+ str(i) +'_academic_year']
            date_of_exam = request.POST['subject'+ str(i) +'_date_of_conduct']
            pr_or_tw = request.POST['subject'+ str(i) +'_pr_or_tw']
            student_count = int(request.POST['subject'+ str(i) +'_student_count'])
            rate = int(request.POST['subject'+ str(i) +'_rate'])
            payable_rate = int(request.POST['subject'+ str(i) +'_payable_rate'])
            
            total = student_count * payable_rate

            data.append([
                i,
                examiner_name,
                category,
                subject_name,
                academic_year,
                date_of_exam,
                pr_or_tw,
                student_count,
                rate,
                payable_rate,
                total
            ])


    # Create the HttpResponse object with the appropriate PDF headers
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="internal_external_bill.pdf"'

    # Create the PDF object, using the HttpResponse object as its "file."
    pdf = canvas.Canvas(response, pagesize=landscape(legal))

    # Set the title
    pdf.setTitle("Exam Payments Report")
    
    col_widths = [30,150,60,170,40,60,30,50,40,90,90,80]

    # Create the table
    table = Table(data,colWidths=col_widths)
    table.setStyle(TableStyle([
        ('SPAN', (0,0), (-1, 0)),  
        ('SPAN', (0, 1), (-1, 1)),

        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 5), (-1, -1), 'Times-Roman'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('WORDWRAP', (0, 0), (-1, -1)),  # Enable text wrapping
        ('FONTSIZE', (0, 0), (-1, -1), 10),  # Reduce font size if necessary
        ('LEFTPADDING', (0, 0), (-1, -1), 3),  # Reduce left padding
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),  # Reduce right padding
    ]))

    page_width, page_height = landscape(legal)

    # Draw the table on the PDF
    table_width, table_height = table.wrapOn(pdf, page_width-20, page_height)
    table.drawOn(pdf, (page_width-table_width)//2, page_height - table_height - 20)


    # Close the PDF object cleanly
    pdf.showPage()
    pdf.save()

    return response
