from django.shortcuts import render
from django.http import HttpResponse

#excel libraries
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side
from io import BytesIO

#pdf libraries
from reportlab.lib.pagesizes import landscape, legal, A4
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from reportlab.platypus import Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle

# Create your views here.
def bills(request):
    return render(request, "select_bill_type.html")

def electricity(request):
    if request.method == 'POST':
        return electricity_excel(request) if request.POST['excel_pdf'] == "excel" else electricity_pdf(request)
    return render(request, "electricity_bills.html")

def electricity_excel(request):
    if request.method == "POST":
        university_name = request.POST['university']
        collage_name = request.POST['collage_name']
        department = request.POST['department']
        time_period = request.POST['time_period']
        subject_count = int(request.POST['subject_count'])
        rate_electricity = int(request.POST['rate_of_electricity'])

        # Create a new Excel workbook and active sheet
        wb = openpyxl.Workbook()
        ws = wb.active

        # Define headers for the table with merged cells
        header_data = [
            ['NAME OF THE COLLEGE : ' + collage_name.upper()],
            ['Statement Showing Claim for use of Computation Facilities, Electricity and allied'],
            [f'EXAM :SE/TE/BE/ME {department} {time_period}'],
            [""],
            ['Sr.No.','Date of Exam','Class','PR/OR/TW/Seminar/Project/Theory',
            'No. of Students', 'Rate','Amount'],
        ]

        # Fill the first header row and merge cells
        for row in header_data:
            ws.append(row)

        # Merging cells for first row as per the requirement
        ws.merge_cells('A1:G1')
        ws.merge_cells('A2:G2')
        ws.merge_cells('A3:G3')
        ws.merge_cells('A4:G4')

        # Append data to the sheet based on subject count
        for i in range(1, subject_count+1):
            subject = request.POST[f"subject{i}"]
            academic_year = request.POST[f"subject{i}_academic_year"]
            exam_date = request.POST[f"subject{i}_date_of_conduct"]
            student_count = int(request.POST[f"subject{i}_student_count"])
            
            amount = rate_electricity * student_count

            # Append data
            ws.append([
                i,
                exam_date,
                academic_year,
                subject,
                student_count,
                rate_electricity,
                amount
            ])
        
        # Apply bold and alignment for header rows
        header_font = Font(bold=True, size=10, name="Times New Roman", color = "002060")
        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in ws.iter_rows(min_row=1, min_col=1, max_col=7):
            for cell in row:
                cell.font = header_font
                cell.alignment = center_alignment

        for row in ws.iter_rows(min_row=1, min_col=1, max_col=7):
            for cell in row:        
                cell.border = thin_border

        ws.column_dimensions["A"].width = 7
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 8
        ws.column_dimensions["D"].width = 40
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 6
        ws.column_dimensions["G"].width = 7

        # Save the workbook to a BytesIO object (in-memory stream)
        file_stream = BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)  # Ensure you reset the stream position

        # Create the response as an Excel file
        response = HttpResponse(file_stream, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="electricity_bill.xlsx"'
        return response

def electricity_pdf(request):
    if request.method == 'POST':
        university_name = request.POST['university']
        collage_name = request.POST['collage_name']
        department = request.POST['department']
        time_period = request.POST['time_period']
        subject_count = int(request.POST['subject_count'])
        rate_electricity = int(request.POST['rate_of_electricity'])
 
        data = [
            ['NAME OF THE COLLEGE : ' + collage_name.upper()],
            ['Statement Showing Claim for use of Computation Facilities, Electricity and allied'],
            [f'EXAM :SE/TE/BE/ME {department} {time_period}'],
            [""],

            ['Sr.No.','Date of Exam','Class', 
            Paragraph('PR/OR/TW/Seminar/Project/Theory',ParagraphStyle(name='CustomStyle',fontSize=10,fontName='Times-Roman',alignment=1)),
            Paragraph('No. of Students', ParagraphStyle(name='CustomStyle',fontSize=10,fontName='Times-Roman',alignment=1)), 
            'Rate','Amount'],
        ]
        
        # Append data to the sheet based on subject count
        for i in range(1, subject_count+1):
            subject = request.POST[f"subject{i}"]
            academic_year = request.POST[f"subject{i}_academic_year"]
            exam_date = request.POST[f"subject{i}_date_of_conduct"]
            student_count = int(request.POST[f"subject{i}_student_count"])
            
            amount = rate_electricity * student_count

            data.append([
                i,
                exam_date,
                academic_year,
                subject,
                student_count,
                rate_electricity,
                amount
            ])


    # Create the HttpResponse object with the appropriate PDF headers
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="electricity_bill.pdf"'

    # Create the PDF object, using the HttpResponse object as its "file."
    pdf = canvas.Canvas(response, pagesize=landscape(A4))

    # Set the title
    pdf.setTitle("Exam Payments Report")
    
    col_widths = [40,70,40,300,50,30,50]

    # Create the table
    table = Table(data,colWidths=col_widths)
    table.setStyle(TableStyle([
        ('SPAN', (0,0), (-1, 0)),  
        ('SPAN', (0, 1), (-1, 1)),
        ('SPAN', (0, 2), (-1, 2)),
        ('SPAN', (0, 3), (-1, 3)),
        ('FONTNAME', (0, 0), (-1, 3), 'Times-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 4), (-1, -1), 'Times-Roman'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('WORDWRAP', (0, 0), (-1, -1)),  # Enable text wrapping
        ('FONTSIZE', (0, 0), (-1, 3), 12),  # Reduce font size if necessary
        ('FONTSIZE', (0, 4), (-1, -1), 10),  # Reduce font size if necessary
        ('LEFTPADDING', (0, 0), (-1, -1), 1),  # Reduce left padding
        ('RIGHTPADDING', (0, 0), (-1, -1), 1),  # Reduce right padding
    ]))

    page_width, page_height = landscape(A4)

    # Draw the table on the PDF
    table_width, table_height = table.wrapOn(pdf, page_width-20, page_height)
    table.drawOn(pdf, (page_width - table_width)//2, page_height - table_height - 40)


    # Close the PDF object cleanly
    pdf.showPage()
    pdf.save()

    return response



def expert_lab_peon(request):
    if request.method == 'POST':
        return expert_lab_peon_excel(request) if request.POST['excel_pdf'] == "excel" else expert_lab_peon_pdf(request)
    return render(request, "expert_lab_peon_bills.html")

def expert_lab_peon_excel(request):
    if request.method == "POST":
        university_name = request.POST['university']
        collage_name = request.POST['college_name']
        time_period = request.POST['time_period']
        subject_count = int(request.POST['subject_count'])

        # Create a new Excel workbook and active sheet
        wb = openpyxl.Workbook()
        ws = wb.active

        # Define headers for the table with merged cells
        header_data = [
            [university_name.upper()],
            ['Statement showing remuneration paid to Lab. Staff / Supervisors/ Expert Asstt., for  Practical  Examinations,'],
            [time_period.upper()],
            ['NAME OF COLLEGE : ' + collage_name.upper()],
            [""],
            ['Sr.No.','Category (Staff/Supervisor/Expert Asstt.) ',
            'Dates of conducts of Exam','No. of staff','Subject',
            'OR/ TW/ PR',"Class",'No. of Students','No. of Batches',
            'Days of','','Exam Rate Rs.', 'Rate of Preparation and cleaning', 
            'Total Amt Rs.'],
            ['','','','','','','','','','Preparation','Cleaning', '']
        ]

        # Fill the first header row and merge cells
        for row in header_data:
            ws.append(row)

        # Merging cells for first row as per the requirement
        ws.merge_cells('A1:N1')
        ws.merge_cells('A2:N2')
        ws.merge_cells('A3:N3')
        ws.merge_cells('A4:N4')
        ws.merge_cells('A5:N5')
        ws.merge_cells('J6:K6')

        # Append data to the sheet based on subject count
        for i in range(1, subject_count+1):
            subject = request.POST[f"subject{i}"]
            category = request.POST[f"subject{i}_category"]
            exam_date = request.POST[f"subject{i}_date_of_conduct"]
            staff_count = int(request.POST[f"subject{i}_staff_count"])
            subject = request.POST[f"subject{i}"]
            pr_or_tw = request.POST[f"subject{i}_pr_or_tw"]
            academic_year = request.POST[f"subject{i}_academic_year"]
            student_count = int(request.POST[f"subject{i}_student_count"])
            batch_count = int(request.POST[f"subject{i}_batch_count"])

            days_of_preparation = int(request.POST[f"subject{i}_days_of_preparation"])
            days_of_cleaning = int(request.POST[f"subject{i}_days_of_cleaning"])
            rate_exam = int(request.POST[f"subject{i}_rate_exam"])
            rate_prep_and_clean = int(request.POST[f"subject{i}_rate_prep_clean"])

            
            total_amount = ((batch_count*rate_exam) + 
            (rate_prep_and_clean* (days_of_preparation + days_of_cleaning))) * staff_count

            # Append data
            ws.append([
                i,
                category,
                exam_date,
                staff_count,
                subject,
                pr_or_tw,
                academic_year,
                student_count,
                batch_count,
                days_of_preparation,
                days_of_cleaning,
                rate_exam,
                rate_prep_and_clean,
                total_amount
            ])
        
        # Apply bold and alignment for header rows
        header_font = Font(bold=True, size=10, name="Times New Roman", color = "002060")
        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        #body style
        for row in ws.iter_rows(min_row=5, min_col=1, max_col=14):
            for cell in row:
                cell.font = header_font
                cell.alignment = center_alignment
        
        #head style
        for row in ws.iter_rows(min_row=1, max_row=4, min_col=1, max_col=14):
            for cell in row:
                cell.font = Font(size=14, bold=True, name="Times New Roman", color = "002060")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for row in ws.iter_rows(min_row=1, min_col=1, max_col=14):
            for cell in row:        
                cell.border = thin_border

        ws.column_dimensions["A"].width = 4
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 5
        ws.column_dimensions["E"].width = 40
        ws.column_dimensions["F"].width = 6
        ws.column_dimensions["G"].width = 6
        ws.column_dimensions["H"].width = 6
        ws.column_dimensions["I"].width = 6
        ws.column_dimensions["J"].width = 6
        ws.column_dimensions["K"].width = 6
        ws.column_dimensions["L"].width = 6
        ws.column_dimensions["M"].width = 6
        ws.column_dimensions["N"].width = 6

        # Save the workbook to a BytesIO object (in-memory stream)
        file_stream = BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)  # Ensure you reset the stream position

        # Create the response as an Excel file
        response = HttpResponse(file_stream, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="expert_lab_peon_bill.xlsx"'
        return response

def expert_lab_peon_pdf(request):
    if request.method == 'POST':
        university_name = request.POST['university']
        collage_name = request.POST['college_name']
        time_period = request.POST['time_period']
        subject_count = int(request.POST['subject_count'])


        data = [
            [university_name.upper()],
            ['Statement showing remuneration paid to Lab. Staff / Supervisors/ Expert Asstt., for  Practical  Examinations,'],
            [time_period.upper()],
            ['NAME OF COLLEGE : ' + collage_name.upper()],
            [""],
            ['Sr.No.',Paragraph('Category (Staff/ Supervisor/ Expert Asstt.) ',ParagraphStyle(name='CustomStyle',fontSize=10,fontName='Times-Roman',alignment=1)),
            Paragraph('Dates of conducts of Exam', ParagraphStyle(name='CustomStyle',fontSize=10,fontName='Times-Roman',alignment=1)),
            'No. of staff','Subject',
            Paragraph('OR/ TW/ PR',ParagraphStyle(name='CustomStyle',fontSize=10,fontName='Times-Roman',alignment=1)),
            "Class",
            Paragraph('No. of Students',ParagraphStyle(name='CustomStyle',fontSize=10,fontName='Times-Roman',alignment=1)),
            Paragraph('No. of Batches',ParagraphStyle(name='CustomStyle',fontSize=10,fontName='Times-Roman',alignment=1)),
            'Days of','','Exam Rate Rs.', 
            Paragraph('Rate of Preparation and cleaning', ParagraphStyle(name='CustomStyle',fontSize=10,fontName='Times-Roman',alignment=1)),
            Paragraph('Total Amt Rs.',ParagraphStyle(name='CustomStyle',fontSize=10,fontName='Times-Roman',alignment=1)),],
            ['','','','','','','','','','Preparation','Cleaning', '']
        ]
        
        # Append data to the sheet based on subject count
        for i in range(1, subject_count+1):
            subject = request.POST[f"subject{i}"]
            category = request.POST[f"subject{i}_category"]
            exam_date = request.POST[f"subject{i}_date_of_conduct"]
            staff_count = int(request.POST[f"subject{i}_staff_count"])
            subject = request.POST[f"subject{i}"]
            pr_or_tw = request.POST[f"subject{i}_pr_or_tw"]
            academic_year = request.POST[f"subject{i}_academic_year"]
            student_count = int(request.POST[f"subject{i}_student_count"])
            batch_count = int(request.POST[f"subject{i}_batch_count"])

            days_of_preparation = int(request.POST[f"subject{i}_days_of_preparation"])
            days_of_cleaning = int(request.POST[f"subject{i}_days_of_cleaning"])
            rate_exam = int(request.POST[f"subject{i}_rate_exam"])
            rate_prep_and_clean = int(request.POST[f"subject{i}_rate_prep_clean"])

            
            total_amount = ((batch_count*rate_exam) + 
            (rate_prep_and_clean* (days_of_preparation + days_of_cleaning))) * staff_count

            # Append data
            data.append([
                i,
                category,
                exam_date,
                staff_count,
                subject,
                pr_or_tw,
                academic_year,
                student_count,
                batch_count,
                days_of_preparation,
                days_of_cleaning,
                rate_exam,
                rate_prep_and_clean,
                total_amount
            ])


    # Create the HttpResponse object with the appropriate PDF headers
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="expert_lab_peon_bill.pdf"'

    # Create the PDF object, using the HttpResponse object as its "file."
    pdf = canvas.Canvas(response, pagesize=landscape(legal))

    # Set the title
    pdf.setTitle("Exam Payments Report")
    
    col_widths = [30,80,60,50,300,40,40,40,40,50,50,60,70,50]

    # Create the table
    table = Table(data,colWidths=col_widths)
    table.setStyle(TableStyle([
        ('SPAN', (0,0), (-1, 0)),  
        ('SPAN', (0, 1), (-1, 1)),
        ('SPAN', (0, 2), (-1, 2)),
        ('SPAN', (0, 3), (-1, 3)),
        ('SPAN', (0, 4), (-1, 4)),
        ('SPAN', (9, 5), (10, 5)),
        ('FONTNAME', (0, 0), (-1, 3), 'Times-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 4), (-1, -1), 'Times-Roman'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('WORDWRAP', (0, 0), (-1, -1)),  # Enable text wrapping
        ('FONTSIZE', (0, 0), (-1, 3), 12),  # Reduce font size if necessary
        ('FONTSIZE', (0, 4), (-1, -1), 10),  # Reduce font size if necessary
        ('LEFTPADDING', (0, 0), (-1, -1), 1),  # Reduce left padding
        ('RIGHTPADDING', (0, 0), (-1, -1), 1),  # Reduce right padding
    ]))

    page_width, page_height = landscape(legal)

    # Draw the table on the PDF
    table_width, table_height = table.wrapOn(pdf, page_width-20, page_height)
    table.drawOn(pdf, (page_width - table_width)//4, page_height - table_height - 40)


    # Close the PDF object cleanly
    pdf.showPage()
    pdf.save()

    return response



def total_lab_and_staff(request):
    if request.method == 'POST':
        return total_lab_and_staff_excel(request)
    return render(request, "total_bill_lab_and_staff.html")

def total_lab_and_staff_excel(request):
    if (request.method == 'POST'):
        university_name = request.POST['university']
        college_name = request.POST['college_name']
        department = request.POST['department']
        time_period = request.POST['time_period']
        staff_count = int(request.POST['staff_count'])

        examiner_fe_sum = 0
        examiner_se_sum = 0
        examiner_te_sum = 0
        examiner_be_sum = 0
        examiner_me_sum = 0
        examiner_total_sum = 0

        expert_fe_sum = 0
        expert_se_sum = 0
        expert_te_sum = 0
        expert_be_sum = 0
        expert_me_sum = 0
        expert_total_sum = 0

        grand_total_sum = 0

        # Create a new Excel workbook and active sheet
        wb = openpyxl.Workbook()
        ws = wb.active

        # Define headers for the table with merged cells
        header_data = [
            [college_name],

            [f"REMUNARATION TO STAFF FOR ORAL & PRACTICAL EXAM {time_period}"],

            [f'DEPARTMENT OF : {department}',""],

            ['INTERNAL EXAMINER'],

            ["Sr. No.", "Name of the Staff", 
            "Rem. To Examiner (Oral & Practical, TW)","","","","","",
            "Rem. To Expert Asstt.", "Rem. To Expert Asstt.","","","","",
            "Grand Total", "Signature" 
            ],

            [
                "", "", 'FE', 'SE', 'TE', 'BE', 'ME', 'TOTAL',
                'FE', 'SE', 'TE', 'BE', 'ME', 'TOTAL', "", ""
            ]
        ]

        # Fill the first header row and merge cells
        for header in header_data:
            ws.append(header)
        
        # Merging cells for first row as per the requirement
        ws.merge_cells('A1:P1')
        ws.merge_cells('A2:P2')
        ws.merge_cells('A3:C3')
        ws.merge_cells('D3:P3')
        ws.merge_cells('A4:P4')
        ws.merge_cells('A5:A6')
        ws.merge_cells('B5:B6')
        
        ws.merge_cells('C5:H5')
        ws.merge_cells('I5:N5')
        ws.merge_cells('O5:O6')
        ws.merge_cells('P5:P6')

        for i in range(1,staff_count+1):
            staff_name = request.POST[f"staff_name{i}"]
            staff_examiner_bill_fe = int(request.POST[f"staff{i}_examiner_FE"])
            staff_examiner_bill_se = int(request.POST[f"staff{i}_examiner_SE"])
            staff_examiner_bill_te = int(request.POST[f"staff{i}_examiner_TE"])
            staff_examiner_bill_be = int(request.POST[f"staff{i}_examiner_BE"])
            staff_examiner_bill_me = int(request.POST[f"staff{i}_examiner_ME"])

            staff_expert_bill_fe = int(request.POST[f"staff{i}_expert_FE"])
            staff_expert_bill_se = int(request.POST[f"staff{i}_expert_SE"])
            staff_expert_bill_te = int(request.POST[f"staff{i}_expert_TE"])
            staff_expert_bill_be = int(request.POST[f"staff{i}_expert_BE"])
            staff_expert_bill_me = int(request.POST[f"staff{i}_expert_ME"])
            
            total_examiner = staff_examiner_bill_fe + \
                            staff_examiner_bill_se + \
                            staff_examiner_bill_te + \
                            staff_examiner_bill_be + \
                            staff_examiner_bill_me


            total_expert = staff_expert_bill_fe + \
                            staff_expert_bill_se + \
                            staff_expert_bill_te + \
                            staff_expert_bill_be + \
                            staff_expert_bill_me

            grand_total = total_examiner + total_expert

            examiner_fe_sum +=  staff_examiner_bill_fe
            examiner_se_sum +=  staff_examiner_bill_se
            examiner_te_sum +=  staff_examiner_bill_te
            examiner_be_sum +=  staff_examiner_bill_be
            examiner_me_sum +=  staff_examiner_bill_me
            examiner_total_sum +=  total_examiner

            expert_fe_sum +=  staff_expert_bill_fe
            expert_se_sum +=  staff_expert_bill_se
            expert_te_sum +=  staff_expert_bill_te
            expert_be_sum +=  staff_expert_bill_be
            expert_me_sum +=  staff_expert_bill_me
            expert_total_sum +=  total_expert

            grand_total_sum +=  grand_total

            #append data
            ws.append([
                i,
                staff_name,
                staff_examiner_bill_fe,
                staff_examiner_bill_se,
                staff_examiner_bill_te,
                staff_examiner_bill_be,
                staff_examiner_bill_me,
                total_examiner,

                staff_expert_bill_fe,
                staff_expert_bill_se,
                staff_expert_bill_te,
                staff_expert_bill_be,
                staff_expert_bill_me,
                total_expert,
                grand_total,
            ])

            # Apply bold and alignment for header rows
            header_font = Font(bold=True,)
            center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Apply font and alignment to header
            for row in ws.iter_rows(min_row=0, max_row=0, min_col=0, max_col=16):
                for cell in row:
                    cell.font = header_font
                    cell.alignment = center_alignment

            for row in ws.iter_rows(min_row=0, min_col=0, max_col=16):
                for cell in row:        
                    cell.border = thin_border

            # Add wrap text to data
            for cell in ws[ws.max_row]:  # Get the last appended row
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 25
        
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 20

        ws.column_dimensions["P"].width = 10

        # Save the workbook to a BytesIO object (in-memory stream)
        file_stream = BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)

        # Create the response as an Excel file
        response = HttpResponse(file_stream, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Total_bill_of_lab_and_staff.xlsx"'

        return response


    return render(request, "statement_of_labstaff_form1.html")



















def int_and_ext_bills(request):
    pass
